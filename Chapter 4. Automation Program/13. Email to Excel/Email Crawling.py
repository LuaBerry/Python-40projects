from tempfile import TemporaryFile
import requests
import re
import os
from openpyxl import load_workbook, Workbook

url = 'https://www.hani.co.kr/arti/economy/it/1045980.html'

headers = {
    'User-Agent': 'Mozilla/5.0',
    'Content-Type': 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w]+\.[\w\.]+', response.text)

results = list(set(results))

dir_path = os.path.dirname(os.path.abspath(__file__))
file_path = dir_path + "\\email.xlsx"

print(results)

try:
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save(file_path)